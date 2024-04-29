import requests
import wget
import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
from fpdf import FPDF



nomecompleto = []

wb = load_workbook('NomeCompleto+registroid.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
    # Obtém o valor da coluna "B"
    nomescompletos = row[3]
    # Adiciona o valor à lista registro_ids
    nomecompleto.append(nomescompletos)

registro_id = []
for cell in ws['C']:
    # Verifique se o valor da célula é um número inteiro
    if isinstance(cell.value, int):
        # Adicione o valor à lista
        registro_id.append(cell.value)

index = 1373

while index < len(registro_id) and index < len(nomecompleto):
    api_url = (f"https://api.eboxdigital.com.br/v1/registros/{registro_id[index]}/documentos")
    token = "bu2AZ8d8Lk9aC0OYA7ZRxA"
    headers = {
        "Content-Type": "application/json",
        "X-Api-Key": "bu2AZ8d8Lk9aC0OYA7ZRxA"
    }
    response = requests.get(api_url, headers=headers, verify=False)
    res = requests.get(api_url, headers=headers, verify=False)
    if res.status_code == 200:
        data = res.json()
        for a in data:
            if a["status"] == "Disponível":
                if a['tipoDocumentoId'] == 4020:
                    link = a['uri']
                    filename = f"{nomecompleto[index]}0.pdf"
                    wget.download(link, filename)
                    file = open(filename, 'rb')
                    reader = PdfReader(file)
                    writer = PdfWriter()

                # Obtenha as últimas 30 páginas
                    for page_number in range(max(0, len(reader.pages) - 30), len(reader.pages)):
                        page = reader.pages[page_number]
                        writer.add_page(page)

                # Salve as últimas 30 páginas em um novo arquivo
                    new_filename = f"{nomecompleto[index]}.pdf"
                    with open(new_filename, 'wb') as output_pdf:
                        writer.write(output_pdf)

                # Feche o arquivo original
                    file.close()

                    os.remove(filename)

        else:
            print(f"Status code: {res.status_code}")
        index += 1  # Incrementa o índice a cada passagem pelo loop




